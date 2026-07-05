"""
Price Comparison page — compare pricing and specs across tracked brands.

Data sources (highest to lowest reliability):
  1. Shopify JSON endpoint — manufacturer MSRP, direct and confirmed
  2. Manufacturer product page HTML — confirmed spec table
  3. Google Shopping HTML — retailer prices + model discovery for comp brands

All data shown is confirmed; any value that cannot be verified is displayed
as "—" rather than guessed.
"""
from __future__ import annotations

from datetime import datetime

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from backend.comp_shopping.comp_shopping_service import CompShoppingService


# ── Worker ────────────────────────────────────────────────────────────────────

class _CompWorker(QThread):
    progress = Signal(str, int, int)   # brand, done, total
    finished = Signal(dict)            # full results dict or {"error": ...}

    def __init__(self, service: CompShoppingService,
                 primary_brand: str, primary_model: str,
                 comp_brands: list[str], keywords: str,
                 retailer_urls: list[str] | None = None):
        super().__init__()
        self.service        = service
        self.primary_brand  = primary_brand
        self.primary_model  = primary_model
        self.comp_brands    = comp_brands
        self.keywords       = keywords
        self.retailer_urls  = retailer_urls or []

    def run(self):
        try:
            result = self.service.run_comparison(
                primary_brand=self.primary_brand,
                primary_model=self.primary_model,
                comp_brands=self.comp_brands,
                keywords=self.keywords,
                retailer_urls=self.retailer_urls,
                progress_callback=lambda b, d, t: self.progress.emit(b, d, t),
            )
            self.finished.emit(result)
        except Exception as exc:
            self.finished.emit({"error": str(exc)})


# ── Page ──────────────────────────────────────────────────────────────────────

class CompShoppingPage(QWidget):

    _CELL_NA = "—"   # displayed for any unconfirmed / missing value

    def __init__(self, app):
        super().__init__()
        self.app     = app
        self.service = CompShoppingService()
        self._worker = None
        self._results: dict = {}
        self._brand_checks: dict[str, QCheckBox] = {}
        self._build_ui()

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout()
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(16)

        # Title row
        title = QLabel("Price Comparison")
        title.setStyleSheet("font-size: 28px; font-weight: 700; color: #111827;")
        subtitle = QLabel(
            "Compare real-time pricing and confirmed product specs across tracked brands."
        )
        subtitle.setStyleSheet("font-size: 13px; color: #6B7280;")
        root.addWidget(title)
        root.addWidget(subtitle)

        # Coming Soon banner
        banner = QFrame()
        banner.setStyleSheet(
            "QFrame { background: #FEF3C7; border: 1px solid #F59E0B; border-radius: 8px; }"
        )
        banner_lay = QHBoxLayout()
        banner_lay.setContentsMargins(14, 10, 14, 10)
        banner_lay.setSpacing(12)
        banner_icon = QLabel("🚧")
        banner_icon.setStyleSheet("font-size: 20px;")
        banner_icon.setFixedWidth(28)
        banner_text = QLabel(
            "<b>Coming Soon</b> — Price Comparison is being redesigned with "
            "AI-assisted product discovery, spec lookup by wattage/fuel type/start type, "
            "and retailer pricing. This page is view-only in the current build."
        )
        banner_text.setWordWrap(True)
        banner_text.setStyleSheet("font-size: 12px; color: #92400E;")
        banner_lay.addWidget(banner_icon)
        banner_lay.addWidget(banner_text, 1)
        banner.setLayout(banner_lay)
        root.addWidget(banner)

        # Main splitter: controls left, results right
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        splitter.addWidget(self._build_controls())
        splitter.addWidget(self._build_results())
        splitter.setSizes([310, 900])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)

        root.addWidget(splitter, 1)
        self.setLayout(root)

    # ── Left control panel ────────────────────────────────────────────────────

    def _build_controls(self) -> QWidget:
        frame = QFrame()
        frame.setFixedWidth(310)
        frame.setStyleSheet(
            "QFrame { background: #F9FAFB; border: 1px solid #E5E7EB; border-radius: 8px; }"
        )
        lay = QVBoxLayout()
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(12)

        # ── Primary product ────────────────────────────────────────────────────
        lay.addWidget(self._section_label("PRIMARY PRODUCT"))

        lay.addWidget(QLabel("Brand"))
        self._primary_brand_cb = QComboBox()
        self._primary_brand_cb.setMinimumHeight(30)
        self._primary_brand_cb.setStyleSheet(self._combo_style())
        lay.addWidget(self._primary_brand_cb)

        lay.addWidget(QLabel("Model Number"))
        self._model_edit = QLineEdit()
        self._model_edit.setPlaceholderText("e.g. FP4000, WGen7500c")
        self._model_edit.setMinimumHeight(30)
        self._model_edit.setStyleSheet(self._input_style())
        lay.addWidget(self._model_edit)

        lay.addWidget(QLabel("Search Keywords"))
        self._keywords_edit = QLineEdit()
        self._keywords_edit.setPlaceholderText("e.g. 7500W dual fuel generator")
        self._keywords_edit.setMinimumHeight(30)
        self._keywords_edit.setStyleSheet(self._input_style())
        self._keywords_edit.setText("generator")
        lay.addWidget(self._keywords_edit)

        tip = QLabel("Keywords refine comparison-brand searches when no model is specified.")
        tip.setWordWrap(True)
        tip.setStyleSheet("font-size: 10px; color: #9CA3AF;")
        lay.addWidget(tip)

        lay.addWidget(QLabel("Retailer Product URLs"))
        self._retailer_urls_edit = QTextEdit()
        self._retailer_urls_edit.setPlaceholderText(
            "Paste product page URLs (one per line):\n"
            "https://www.lowes.com/pd/...\n"
            "https://www.homedepot.com/p/..."
        )
        self._retailer_urls_edit.setMinimumHeight(70)
        self._retailer_urls_edit.setMaximumHeight(90)
        self._retailer_urls_edit.setStyleSheet(
            "QTextEdit { background: white; border: 1px solid #D1D5DB; "
            "border-radius: 5px; padding: 4px 8px; font-size: 10px; "
            "color: #374151; }"
            "QTextEdit:focus { border-color: #0B84FF; }"
        )
        lay.addWidget(self._retailer_urls_edit)

        url_tip = QLabel(
            "Atlas will scrape these pages for current retail prices. "
            "Paste the exact product page URL from each retailer."
        )
        url_tip.setWordWrap(True)
        url_tip.setStyleSheet("font-size: 10px; color: #9CA3AF;")
        lay.addWidget(url_tip)

        lay.addWidget(self._divider())

        # ── Comparison brands ──────────────────────────────────────────────────
        lay.addWidget(self._section_label("COMPARE AGAINST"))

        btn_row = QHBoxLayout()
        self._top_mentions_btn = QPushButton("Select Top Mentions")
        self._top_mentions_btn.setToolTip(
            "Auto-select top brands from Visibility data"
        )
        self._top_mentions_btn.clicked.connect(self._select_top_mentions)
        self._top_mentions_btn.setStyleSheet(self._secondary_btn_style())

        clear_btn = QPushButton("Clear")
        clear_btn.clicked.connect(self._clear_comparison)
        clear_btn.setStyleSheet(self._secondary_btn_style())
        clear_btn.setFixedWidth(56)

        btn_row.addWidget(self._top_mentions_btn)
        btn_row.addWidget(clear_btn)
        lay.addLayout(btn_row)

        # Brand checkbox grid inside scroll area
        self._brand_scroll = QScrollArea()
        self._brand_scroll.setWidgetResizable(True)
        self._brand_scroll.setFrameShape(QFrame.NoFrame)
        self._brand_scroll.setStyleSheet("background: transparent;")
        self._brand_scroll.setMinimumHeight(220)

        self._brand_grid_widget = QWidget()
        self._brand_grid_widget.setStyleSheet("background: transparent;")
        self._brand_grid = QGridLayout(self._brand_grid_widget)
        self._brand_grid.setContentsMargins(0, 0, 0, 0)
        self._brand_grid.setSpacing(4)
        self._brand_scroll.setWidget(self._brand_grid_widget)
        lay.addWidget(self._brand_scroll, 1)

        lay.addWidget(self._divider())

        # ── Run button ─────────────────────────────────────────────────────────
        self._run_btn = QPushButton("Run Comparison")
        self._run_btn.setMinimumHeight(38)
        self._run_btn.setCursor(Qt.PointingHandCursor)
        self._run_btn.setStyleSheet(
            "QPushButton { font-size: 13px; font-weight: 700; color: white; "
            "background: #0B84FF; border: none; border-radius: 6px; padding: 6px 16px; }"
            "QPushButton:hover { background: #0056CC; }"
            "QPushButton:pressed { background: #003D99; }"
            "QPushButton:disabled { background: #9CA3AF; }"
        )
        self._run_btn.clicked.connect(self._run_comparison)
        lay.addWidget(self._run_btn)

        self._status_lbl = QLabel("")
        self._status_lbl.setWordWrap(True)
        self._status_lbl.setStyleSheet("font-size: 10px; color: #6B7280;")
        lay.addWidget(self._status_lbl)

        frame.setLayout(lay)
        frame.setEnabled(False)
        return frame

    # ── Right results panel ───────────────────────────────────────────────────

    def _build_results(self) -> QWidget:
        widget = QWidget()
        lay    = QVBoxLayout()
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(8)

        # Toolbar row
        self._last_run_lbl = QLabel("No comparison run yet.")
        self._last_run_lbl.setStyleSheet("font-size: 11px; color: #6B7280;")

        export_btn = QPushButton("Export Excel")
        export_btn.setFixedHeight(28)
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.setStyleSheet(
            "QPushButton { font-size: 11px; font-weight: 600; color: #0B84FF; "
            "background: white; border: 1.5px solid #0B84FF; border-radius: 5px; "
            "padding: 2px 12px; }"
            "QPushButton:hover { background: #EFF6FF; }"
        )
        export_btn.clicked.connect(self._export_excel)

        toolbar = QHBoxLayout()
        toolbar.addWidget(self._last_run_lbl, 1)
        toolbar.addWidget(export_btn)
        lay.addLayout(toolbar)

        # Progress bar (hidden until run starts)
        self._progress_lbl = QLabel("")
        self._progress_lbl.setStyleSheet("font-size: 11px; color: #374151;")
        self._progress_lbl.hide()
        lay.addWidget(self._progress_lbl)

        # Result tabs
        self._tabs = QTabWidget()
        self._tabs.setStyleSheet(
            "QTabBar::tab { padding: 6px 16px; font-size: 12px; }"
            "QTabBar::tab:selected { font-weight: 600; color: #0B84FF; "
            "border-bottom: 2px solid #0B84FF; }"
        )

        self._price_table  = self._make_table([
            "Brand", "Model",
            "Mfr. MSRP", "Best Price", "Best @ Retailer",
            "Prev Best", "Change", "Updated",
        ])
        self._spec_table   = self._make_table(["Spec"])  # columns added dynamically
        self._status_table = self._make_table(["Brand", "Model/Search", "Results", "Status"])

        self._tabs.addTab(self._wrap_table(self._price_table), "Price Comparison")
        self._tabs.addTab(self._wrap_table(self._spec_table),  "Spec Comparison")
        self._tabs.addTab(self._wrap_table(self._status_table), "Data Status")

        lay.addWidget(self._tabs, 1)

        # Data quality note
        note = QLabel(
            "Sources: Manufacturer Direct (Shopify JSON) → Manufacturer Page (specs) → Google Shopping (retailer prices / model discovery).  "
            "— = data not found on any source.  No values are inferred or estimated."
        )
        note.setStyleSheet("font-size: 10px; color: #9CA3AF;")
        note.setWordWrap(True)
        lay.addWidget(note)

        widget.setLayout(lay)
        return widget

    # ── Populate brands ───────────────────────────────────────────────────────

    def refresh(self):
        self._populate_brands()

    def _populate_brands(self):
        try:
            from backend.services.knowledge_service import KnowledgeService
            brands = sorted(KnowledgeService().get_brands())
        except Exception:
            brands = []

        # Primary brand combo
        current = self._primary_brand_cb.currentText()
        self._primary_brand_cb.clear()
        self._primary_brand_cb.addItems(brands)
        if current in brands:
            self._primary_brand_cb.setCurrentText(current)

        # Comparison checkboxes
        while self._brand_grid.count():
            item = self._brand_grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._brand_checks.clear()

        for i, brand in enumerate(brands):
            cb = QCheckBox(brand)
            cb.setStyleSheet("font-size: 11px;")
            self._brand_grid.addWidget(cb, i // 2, i % 2)
            self._brand_checks[brand] = cb

    # ── Top Mentions auto-select ───────────────────────────────────────────────

    def _select_top_mentions(self):
        try:
            from backend.visibility.visibility_service import VisibilityService
            svc    = VisibilityService(self.app.provider_manager,
                                       target_brand=self.app.get_target_brand())
            counts = svc.analytics_summary().get("brand_counts", {})
            if not counts:
                self._status_lbl.setText("No visibility data yet — run a collection first.")
                return
            primary = self._primary_brand_cb.currentText()
            top10   = [b for b, _ in
                       sorted(counts.items(), key=lambda x: -x[1])
                       if b != primary][:10]
            for brand, cb in self._brand_checks.items():
                cb.setChecked(brand in top10)
            self._status_lbl.setText(
                f"Selected {len(top10)} top brands from visibility data."
            )
        except Exception as exc:
            self._status_lbl.setText(f"Could not load visibility data: {exc}")

    def _clear_comparison(self):
        for cb in self._brand_checks.values():
            cb.setChecked(False)

    # ── Run ───────────────────────────────────────────────────────────────────

    def _run_comparison(self):
        primary_brand = self._primary_brand_cb.currentText().strip()
        primary_model = self._model_edit.text().strip()
        keywords      = self._keywords_edit.text().strip() or "generator"
        comp_brands   = [b for b, cb in self._brand_checks.items()
                         if cb.isChecked() and b != primary_brand]
        retailer_urls = [
            u.strip() for u in self._retailer_urls_edit.toPlainText().splitlines()
            if u.strip().startswith("http")
        ]

        if not primary_brand:
            QMessageBox.warning(self, "Missing Input",
                                "Please select a primary brand.")
            return
        if not comp_brands:
            QMessageBox.warning(self, "Missing Input",
                                "Select at least one comparison brand.")
            return

        self._run_btn.setEnabled(False)
        self._run_btn.setText("Running…")
        self._progress_lbl.show()
        self._progress_lbl.setText("Starting…")

        self._worker = _CompWorker(
            self.service, primary_brand, primary_model,
            comp_brands, keywords, retailer_urls=retailer_urls,
        )
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_finished)
        self._worker.start()

    def _on_progress(self, brand: str, done: int, total: int):
        self._progress_lbl.setText(
            f"Searching {brand}… ({done}/{total})"
        )

    def _on_finished(self, result: dict):
        self._run_btn.setEnabled(True)
        self._run_btn.setText("Run Comparison")
        self._progress_lbl.hide()

        if "error" in result:
            QMessageBox.critical(self, "Error",
                                 f"Comparison failed:\n{result['error']}")
            return

        self._results = result
        self._last_run_lbl.setText(
            f"Last run: {datetime.now().strftime('%b %d, %Y  %H:%M')}"
        )
        self._populate_price_table(result["brands"])
        self._populate_spec_table(result["brands"])
        self._populate_status_table(result["brands"])

    # ── Table population ──────────────────────────────────────────────────────

    def _populate_price_table(self, brands: list[dict]):
        """
        One row per brand/model combo.

        Columns: Brand | Model | Mfr. MSRP | Best Price | Best @ Retailer
                 | Prev Best | Change | Updated

        MSRP = price from manufacturer Shopify direct.
        Best Price = lowest price from any retailer (incl. user-supplied URLs).
        When multiple models are present for a brand (DuroMax etc.), each gets its own row.
        """
        tbl = self._price_table
        tbl.setRowCount(0)
        today = datetime.now().strftime("%Y-%m-%d")

        for entry in brands:
            brand       = entry["brand"]
            entry_model = entry["model"] or entry["search_q"]
            prices      = entry.get("prices", [])

            if not prices:
                row = tbl.rowCount()
                tbl.insertRow(row)
                self._set_item(tbl, row, 0, brand)
                self._set_item(tbl, row, 1, entry_model)
                na = QTableWidgetItem("No data found")
                na.setForeground(Qt.gray)
                tbl.setItem(row, 2, na)
                for col in range(3, 8):
                    tbl.setItem(row, col, QTableWidgetItem(self._CELL_NA))
                continue

            # Group prices by model: {model_key: [price_dicts]}
            model_groups: dict[str, list[dict]] = {}
            for p in prices:
                key = p.get("model_extracted") or entry_model
                model_groups.setdefault(key, []).append(p)

            for model_key, model_prices in model_groups.items():
                # Manufacturer MSRP = shopify_direct price
                msrp_entries = [p for p in model_prices if p.get("method") == "shopify_direct"]
                msrp_price   = min((p["price"] for p in msrp_entries), default=None)

                # Best retail = lowest price NOT from manufacturer
                retail_entries = [p for p in model_prices if p.get("method") != "shopify_direct"]
                # Also include user-supplied retailer_direct prices
                retail_entries += [p for p in model_prices if p.get("method") == "retailer_direct"]
                # De-dup (retailer_direct might appear in both lists if method check above is off)
                seen = set()
                unique_retail = []
                for p in retail_entries:
                    if id(p) not in seen:
                        seen.add(id(p))
                        unique_retail.append(p)

                best_retail = None
                best_retailer = self._CELL_NA
                best_url = ""
                if unique_retail:
                    best_entry  = min(unique_retail, key=lambda x: x["price"])
                    best_retail = best_entry["price"]
                    best_retailer = best_entry.get("retailer", self._CELL_NA)
                    best_url    = best_entry.get("url", "")

                # If no dedicated retail prices, use lowest MSRP as best price
                best_price = best_retail if best_retail is not None else msrp_price
                best_retail_label = best_retailer if best_retail is not None else (
                    "Mfr. Direct" if msrp_price else self._CELL_NA
                )

                # Previous best price (from any source)
                all_prev = [p.get("prev_price") for p in model_prices if p.get("prev_price")]
                prev_best = min(all_prev) if all_prev else None

                chg = None
                if best_price and prev_best and prev_best > 0:
                    chg = round((best_price - prev_best) / prev_best * 100, 1)

                msrp_str  = f"${msrp_price:,.2f}" if msrp_price else self._CELL_NA
                best_str  = f"${best_price:,.2f}" if best_price else self._CELL_NA
                prev_str  = f"${prev_best:,.2f}" if prev_best else self._CELL_NA
                chg_str   = (f"{'▼' if chg < 0 else '▲'} {abs(chg):.1f}%"
                             if chg is not None else self._CELL_NA)

                row = tbl.rowCount()
                tbl.insertRow(row)

                vals = [brand, model_key, msrp_str, best_str,
                        best_retail_label, prev_str, chg_str, today]
                for col, val in enumerate(vals):
                    item = QTableWidgetItem(str(val))
                    item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                    if col == 6 and chg is not None:
                        item.setForeground(Qt.darkGreen if chg < 0 else Qt.red)
                    if col == 3 and best_url:
                        item.setToolTip(best_url)
                    # Highlight best price when it's better than MSRP
                    if col == 3 and best_retail is not None and msrp_price and best_retail < msrp_price:
                        item.setForeground(Qt.darkGreen)
                    tbl.setItem(row, col, item)

        tbl.resizeColumnsToContents()
        tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)

    def _populate_spec_table(self, brands: list[dict]):
        tbl = self._spec_table

        # Collect all confirmed spec names across all brands
        all_specs: list[str] = []
        seen: set[str] = set()
        for entry in brands:
            for spec in entry.get("specs", {}):
                if spec not in seen:
                    seen.add(spec)
                    all_specs.append(spec)

        if not all_specs:
            tbl.setColumnCount(1)
            tbl.setHorizontalHeaderLabels(["Spec"])
            tbl.setRowCount(1)
            tbl.setItem(0, 0, QTableWidgetItem(
                "No spec data found. Enter model numbers and re-run."
            ))
            return

        # Build columns: Spec | Brand1 | Brand2 | ...
        col_headers = ["Spec"] + [
            f"{e['brand']}\n{e['model']}" if e["model"] else e["brand"]
            for e in brands
        ]
        tbl.setColumnCount(len(col_headers))
        tbl.setHorizontalHeaderLabels(col_headers)
        tbl.setRowCount(len(all_specs))

        for row, spec in enumerate(all_specs):
            tbl.setItem(row, 0, QTableWidgetItem(spec))
            for col, entry in enumerate(brands, 1):
                val = entry.get("specs", {}).get(spec, self._CELL_NA)
                item = QTableWidgetItem(val)
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                if val == self._CELL_NA:
                    item.setForeground(Qt.gray)
                tbl.setItem(row, col, item)

        tbl.resizeColumnsToContents()
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)

    def _populate_status_table(self, brands: list[dict]):
        tbl = self._status_table
        tbl.setRowCount(0)

        for entry in brands:
            row = tbl.rowCount()
            tbl.insertRow(row)
            n_prices = len(entry["prices"])
            n_specs  = len(entry.get("specs", {}))
            status   = entry.get("status", "ok")

            status_str = (
                f"✓ {n_prices} price(s), {n_specs} spec(s) confirmed"
                if status == "ok" and n_prices > 0
                else f"⚠ {status}"
            )

            tbl.setItem(row, 0, QTableWidgetItem(entry["brand"]))
            tbl.setItem(row, 1, QTableWidgetItem(
                entry["model"] or entry["search_q"]
            ))
            tbl.setItem(row, 2, QTableWidgetItem(
                f"{n_prices} prices | {n_specs} specs"
            ))
            status_item = QTableWidgetItem(status_str)
            status_item.setForeground(
                Qt.darkGreen if "✓" in status_str else Qt.darkYellow
            )
            tbl.setItem(row, 3, status_item)

        tbl.resizeColumnsToContents()

    # ── Export ────────────────────────────────────────────────────────────────

    def _export_excel(self):
        if not self._results:
            QMessageBox.information(self, "No Data",
                                    "Run a comparison first.")
            return
        import os
        from PySide6.QtWidgets import QFileDialog
        default = os.path.join(
            os.path.expanduser("~"), "Downloads",
            f"Atlas_CompShop_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", default,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if not path:
            return
        try:
            self._write_excel(path)
            import subprocess
            os.startfile(path)
        except Exception as exc:
            QMessageBox.critical(self, "Export Failed", str(exc))

    def _write_excel(self, path: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        NAVY  = "1E3A5F"
        WHITE = "FFFFFF"
        ALT   = "F3F4F6"

        def hdr_style(cell):
            cell.font      = Font(bold=True, color=WHITE, name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── Price sheet ───────────────────────────────────────────────────────
        ws1 = wb.create_sheet("Price Comparison")
        headers = ["Brand", "Model", "Mfr. MSRP", "Best Price",
                   "Best @ Retailer", "Prev Best", "Change %", "All Prices (detail)"]
        for c, h in enumerate(headers, 1):
            hdr_style(ws1.cell(1, c, h))

        row = 2
        for entry in self._results.get("brands", []):
            em = entry["model"] or entry["search_q"]
            prices = entry.get("prices", [])
            # Group by model key
            groups: dict[str, list] = {}
            for p in prices:
                key = p.get("model_extracted") or em
                groups.setdefault(key, []).append(p)

            for mk, mprices in groups.items():
                msrp_p  = min((p["price"] for p in mprices if p.get("method") == "shopify_direct"), default=None)
                retail  = [p for p in mprices if p.get("method") != "shopify_direct"]
                best_p  = min((p["price"] for p in retail), default=None)
                best_r  = next((p.get("retailer","") for p in retail if p["price"] == best_p), "") if best_p else ""
                prev_p  = min((p["prev_price"] for p in mprices if p.get("prev_price")), default=None)
                chg = (round((best_p - prev_p)/prev_p*100,1) if best_p and prev_p else None)
                detail = "; ".join(f"{p.get('retailer','')} ${p['price']}" for p in mprices)

                ws1.cell(row, 1, entry["brand"])
                ws1.cell(row, 2, mk)
                ws1.cell(row, 3, msrp_p)
                ws1.cell(row, 4, best_p or msrp_p)
                ws1.cell(row, 5, best_r or ("Mfr. Direct" if msrp_p else ""))
                ws1.cell(row, 6, prev_p)
                ws1.cell(row, 7, f"{chg:+.1f}%" if chg is not None else "")
                ws1.cell(row, 8, detail)
                row += 1

        # ── Spec sheet ────────────────────────────────────────────────────────
        ws2    = wb.create_sheet("Spec Comparison")
        brands = self._results.get("brands", [])
        all_specs = []
        seen_s: set[str] = set()
        for e in brands:
            for s in e.get("specs", {}):
                if s not in seen_s:
                    seen_s.add(s); all_specs.append(s)

        spec_cols = ["Spec"] + [
            f"{e['brand']} {e['model']}".strip() for e in brands
        ]
        for c, h in enumerate(spec_cols, 1):
            hdr_style(ws2.cell(1, c, h))
        for r, spec in enumerate(all_specs, 2):
            ws2.cell(r, 1, spec)
            for c, entry in enumerate(brands, 2):
                ws2.cell(r, c, entry.get("specs", {}).get(spec, ""))

        wb.save(path)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _set_item(self, tbl: QTableWidget, row: int, col: int, val: str):
        item = QTableWidgetItem(str(val))
        item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        tbl.setItem(row, col, item)

    def _make_table(self, headers: list[str]) -> QTableWidget:
        tbl = QTableWidget(0, len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.setAlternatingRowColors(True)
        tbl.verticalHeader().hide()
        tbl.horizontalHeader().setStretchLastSection(True)
        tbl.setStyleSheet(
            "QTableWidget { gridline-color: #E5E7EB; font-size: 12px; }"
            "QHeaderView::section { background: #1E3A5F; color: white; "
            "font-weight: 600; font-size: 11px; padding: 6px 8px; "
            "border: none; border-right: 1px solid #2D5A8E; }"
            "QTableWidget::item:alternate { background: #F9FAFB; }"
        )
        return tbl

    def _wrap_table(self, tbl: QTableWidget) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout()
        lay.setContentsMargins(0, 8, 0, 0)
        lay.addWidget(tbl)
        w.setLayout(lay)
        return w

    def _section_label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            "font-size: 10px; font-weight: 700; color: #6B7280; letter-spacing: 0.5px;"
        )
        return lbl

    def _divider(self) -> QFrame:
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("color: #E5E7EB;")
        return line

    def _combo_style(self) -> str:
        return (
            "QComboBox { background: white; border: 1px solid #D1D5DB; "
            "border-radius: 5px; padding: 4px 8px; font-size: 12px; }"
            "QComboBox:focus { border-color: #0B84FF; }"
        )

    def _input_style(self) -> str:
        return (
            "QLineEdit { background: white; border: 1px solid #D1D5DB; "
            "border-radius: 5px; padding: 4px 8px; font-size: 12px; }"
            "QLineEdit:focus { border-color: #0B84FF; }"
        )

    def _secondary_btn_style(self) -> str:
        return (
            "QPushButton { font-size: 11px; font-weight: 600; color: #374151; "
            "background: white; border: 1px solid #D1D5DB; border-radius: 5px; "
            "padding: 4px 10px; }"
            "QPushButton:hover { background: #F3F4F6; }"
        )
