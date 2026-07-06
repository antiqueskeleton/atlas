"""
Excel report generator for Atlas AI Visibility Intelligence reports.
Uses openpyxl for .xlsx output with multi-sheet layout.
"""
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class VisibilityExcelReport:
    """
    Generate a multi-sheet .xlsx from visibility analytics + raw response data.

    Sheets:
        Summary              — KPI overview + provider/prompt-family scores
        Brand Analytics      — mention rate, position breakdown per brand
        Provider Scores      — brand × provider mention pivot
        Feature Intelligence — feature totals + brand × feature pivot
        Channel Intelligence — channel mentions + competitive gap analysis
        Raw Responses        — full response log (capped at 10,000 rows)

    Usage:
        rpt = VisibilityExcelReport(analytics, runs, stats,
                                    target_brand="Firman",
                                    raw_responses=repo.list_responses())
        rpt.generate("/path/to/report.xlsx")
    """

    NAVY    = "1E3A5F"
    LIGHT   = "D1D5DB"
    ROW_A   = "F3F4F6"
    TARGET  = "DBEAFE"   # blue tint for target-brand rows
    WHITE   = "FFFFFF"
    GRAY_FG = "6B7280"

    def __init__(self, analytics: dict, runs: list, stats: dict,
                 target_brand: str = "", raw_responses: list = None):
        self.analytics     = analytics
        self.runs          = runs
        self.stats         = stats
        self.target_brand  = target_brand or analytics.get("target_brand", "Brand")
        self.raw_responses = raw_responses or []
        self.generated_at  = datetime.now()

    # ── Public ─────────────────────────────────────────────────────────────────

    def generate(self, output_path: str) -> None:
        wb = Workbook()
        wb.remove(wb.active)

        self._sheet_summary(wb)
        self._sheet_brand_analytics(wb)
        self._sheet_provider_scores(wb)
        self._sheet_feature_intelligence(wb)
        self._sheet_channel_intelligence(wb)
        if self.raw_responses:
            self._sheet_raw_responses(wb)

        wb.save(output_path)

    # ── Style helpers ──────────────────────────────────────────────────────────

    def _hdr_fill(self):
        return PatternFill("solid", fgColor=self.NAVY)

    def _row_fill(self, row: int, highlight: bool):
        if highlight:
            return PatternFill("solid", fgColor=self.TARGET)
        return PatternFill("solid", fgColor=(self.ROW_A if row % 2 == 0 else self.WHITE))

    def _thin_border(self):
        s = Side(style="thin", color=self.LIGHT)
        return Border(left=s, right=s, top=s, bottom=s)

    def _write_header(self, ws, row: int, cols: list):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.fill      = self._hdr_fill()
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = self._thin_border()

    def _write_row(self, ws, row: int, vals: list, highlight: bool = False):
        fill = self._row_fill(row, highlight)
        bold = highlight
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font      = Font(bold=bold, size=10, name="Calibri", color="111827")
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = self._thin_border()

    def _section_label(self, ws, row: int, text: str):
        ws.cell(row=row, column=1, value=text).font = Font(
            bold=True, size=11, color=self.NAVY, name="Calibri"
        )

    def _auto_width(self, ws, min_w: int = 8, max_w: int = 50):
        for col_cells in ws.columns:
            col = get_column_letter(col_cells[0].column)
            w   = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col].width = max(min_w, min(w + 2, max_w))

    # ── Sheet: Summary ─────────────────────────────────────────────────────────

    def _sheet_summary(self, wb: Workbook):
        ws = wb.create_sheet("Summary")
        tb = self.target_brand
        a  = self.analytics

        ws.cell(1, 1, f"Atlas AI Visibility Report — {tb}").font = Font(
            bold=True, size=14, color=self.NAVY, name="Calibri"
        )
        ws.cell(2, 1,
                f"Generated {self.generated_at.strftime('%B %d, %Y  %H:%M')}").font = Font(
            italic=True, size=9, color=self.GRAY_FG, name="Calibri"
        )

        brand_cnts = a.get("brand_counts", {})
        sorted_b   = sorted(brand_cnts.items(), key=lambda x: -x[1])
        rank       = next((i + 1 for i, (b, _) in enumerate(sorted_b)
                           if b == tb), None)
        # total tracked brands, not just ones with >=1 mention in this dataset —
        # same fix as #48 already applied to visibility_page.py; this report
        # generator had its own separate copy of the old, wrong denominator.
        total_tracked = a.get("total_tracked_brands", len(sorted_b))

        self._section_label(ws, 4, "KEY PERFORMANCE INDICATORS")
        self._write_header(ws, 5, ["Metric", "Value"])
        kpis = [
            ("Visibility Score",     f"{a.get('target_visibility_score', 0)}%"),
            ("Mention Rank",         f"#{rank} of {total_tracked}" if rank else "—"),
            ("Total Responses",      f"{a.get('total_responses', 0):,}"),
            ("Brands Tracked",       str(total_tracked)),
            ("Collection Runs",      str(self.stats.get("runs", 0))),
            ("Providers",            str(self.stats.get("providers", 0))),
            ("Prompt Families",      str(self.stats.get("families", 0))),
        ]
        for i, (k, v) in enumerate(kpis, 6):
            self._write_row(ws, i, [k, v])

        row = len(kpis) + 8

        prov_scores = a.get("provider_visibility_scores", {})
        if prov_scores:
            self._section_label(ws, row, "VISIBILITY SCORE BY PROVIDER")
            row += 1
            self._write_header(ws, row, ["Provider", f"{tb} Visibility Score"])
            row += 1
            for prov, sc in sorted(prov_scores.items(), key=lambda x: -x[1]):
                self._write_row(ws, row, [prov.capitalize(), f"{sc}%"])
                row += 1
            row += 1

        prompt_scores = a.get("prompt_set_visibility_scores", {})
        if prompt_scores:
            self._section_label(ws, row, "VISIBILITY SCORE BY PROMPT FAMILY")
            row += 1
            self._write_header(ws, row, ["Prompt Family", f"{tb} Visibility Score"])
            row += 1
            for ps, sc in sorted(prompt_scores.items(), key=lambda x: -x[1]):
                self._write_row(ws, row, [ps, f"{sc}%"])
                row += 1

        ws.freeze_panes = "A6"
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 24

    # ── Sheet: Brand Analytics ─────────────────────────────────────────────────

    def _sheet_brand_analytics(self, wb: Workbook):
        ws  = wb.create_sheet("Brand Analytics")
        tb  = self.target_brand
        a   = self.analytics

        brand_cnts  = a.get("brand_counts", {})
        total       = max(a.get("total_responses", 1), 1)
        first_share = a.get("first_mention_share", {})
        pos_counts  = a.get("brand_position_counts", {})
        pos_share   = a.get("brand_position_share", {})
        positions   = sorted(pos_counts.keys()) if pos_counts else []
        neg_counts  = a.get("negative_brand_counts", {})
        neg_rate    = a.get("brand_negative_rate", {})

        headers = (
            ["Brand", "Mentions", "Mention Rate %", "First-Mention %",
             "Negative Mentions", "Negative %"]
            + [f"#{p} Count" for p in positions]
            + [f"#{p} Share %" for p in positions]
        )
        self._write_header(ws, 1, headers)

        for row, (brand, cnt) in enumerate(
            sorted(brand_cnts.items(), key=lambda x: -x[1]), 2
        ):
            self._write_row(ws, row, [
                brand,
                cnt,
                round(cnt / total * 100, 1),
                first_share.get(brand, 0),
                neg_counts.get(brand, 0),
                neg_rate.get(brand, 0),
            ] + [pos_counts.get(p, {}).get(brand, 0) for p in positions]
              + [pos_share.get(p, {}).get(brand, 0) for p in positions],
                highlight=(brand == tb))

        ws.freeze_panes = "A2"
        self._auto_width(ws)
        ws.column_dimensions["A"].width = 22

    # ── Sheet: Provider Scores ─────────────────────────────────────────────────

    def _sheet_provider_scores(self, wb: Workbook):
        ws = wb.create_sheet("Provider Scores")
        tb = self.target_brand
        a  = self.analytics

        prov_brand = a.get("provider_brand_counts", {})
        if not prov_brand:
            return

        providers  = sorted(prov_brand.keys())
        all_brands = sorted(
            {b for pb in prov_brand.values() for b in pb},
            key=lambda b: -sum(prov_brand[p].get(b, 0) for p in providers),
        )

        self._write_header(ws, 1, ["Brand"] + [p.capitalize() for p in providers] + ["Total"])
        for row, brand in enumerate(all_brands, 2):
            counts = [prov_brand[p].get(brand, 0) for p in providers]
            self._write_row(ws, row, [brand] + counts + [sum(counts)],
                            highlight=(brand == tb))

        ws.freeze_panes = "A2"
        self._auto_width(ws)
        ws.column_dimensions["A"].width = 22

    # ── Sheet: Feature Intelligence ────────────────────────────────────────────

    def _sheet_feature_intelligence(self, wb: Workbook):
        ws = wb.create_sheet("Feature Intelligence")
        tb = self.target_brand
        a  = self.analytics

        feature_cnts = a.get("feature_counts", {})
        feat_brand   = a.get("feature_brand_counts", {})
        if not feature_cnts:
            return

        all_features = sorted(feature_cnts.keys())
        all_brands   = sorted(
            {b for fb in feat_brand.values() for b in fb},
            key=lambda b: -sum(feat_brand.get(f, {}).get(b, 0) for f in all_features),
        )

        # Feature totals
        self._section_label(ws, 1, "Feature Mention Totals")
        self._write_header(ws, 2, ["Feature", "Total Mentions"])
        for row, (feat, cnt) in enumerate(
            sorted(feature_cnts.items(), key=lambda x: -x[1]), 3
        ):
            self._write_row(ws, row, [feat, cnt])

        # Brand × Feature pivot
        pivot_row = len(feature_cnts) + 5
        self._section_label(ws, pivot_row, "Brand × Feature Pivot")
        pivot_row += 1
        self._write_header(ws, pivot_row, ["Brand"] + all_features + ["Total"])
        for row, brand in enumerate(all_brands, pivot_row + 1):
            counts = [feat_brand.get(f, {}).get(brand, 0) for f in all_features]
            self._write_row(ws, row, [brand] + counts + [sum(counts)],
                            highlight=(brand == tb))

        ws.freeze_panes = "A3"
        self._auto_width(ws)
        ws.column_dimensions["A"].width = 22

    # ── Sheet: Channel Intelligence ────────────────────────────────────────────

    def _sheet_channel_intelligence(self, wb: Workbook):
        ws = wb.create_sheet("Channel Intelligence")
        tb = self.target_brand
        a  = self.analytics

        channel_cnts = a.get("channel_counts", {})
        ch_brand     = a.get("channel_brand_counts", {})
        gaps         = a.get("target_channel_gap", [])

        row = 1
        if channel_cnts:
            self._section_label(ws, row, "Channel Mention Frequency")
            row += 1
            self._write_header(ws, row, ["Channel", "Total Mentions", "Top Brands"])
            row += 1
            for ch, cnt in sorted(channel_cnts.items(), key=lambda x: -x[1]):
                top_b = ", ".join(
                    f"{b} ({c})"
                    for b, c in sorted(ch_brand.get(ch, {}).items(),
                                       key=lambda x: -x[1])[:6]
                )
                self._write_row(ws, row, [ch, cnt, top_b])
                row += 1
            row += 1

        if gaps:
            self._section_label(ws, row, f"{tb} Channel Gaps")
            row += 1
            self._write_header(ws, row, [
                "Channel", tb, "Top Competitor", "Their Count",
                "Gap", "Lead Factor", "Total Competitor Mentions",
            ])
            row += 1
            for g in gaps:
                target_n = g.get("target_count", 0)
                comp_n   = g["top_competitor_count"]
                factor   = (f"{comp_n / target_n:.1f}×"
                            if target_n > 0 else f"{comp_n}×")
                self._write_row(ws, row, [
                    g["channel"],
                    target_n,
                    g["top_competitor"],
                    comp_n,
                    comp_n - target_n,
                    factor,
                    g.get("total_competitor_mentions", ""),
                ])
                row += 1

        ws.freeze_panes = "A3"
        self._auto_width(ws)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["C"].width = 45   # top brands / competitor name

    # ── Sheet: Raw Responses ───────────────────────────────────────────────────

    def _sheet_raw_responses(self, wb: Workbook):
        ws = wb.create_sheet("Raw Responses")

        self._write_header(ws, 1, [
            "ID", "Run ID", "Provider", "Model",
            "Prompt Family", "Prompt", "Response", "Collected At",
        ])

        for row, resp in enumerate(self.raw_responses[:10_000], 2):
            # (id, run_id, provider, model, prompt, response, collected_at, family_display)
            self._write_row(ws, row, [
                resp[0],   # id
                resp[1],   # run_id
                resp[2],   # provider
                resp[3],   # model
                resp[7],   # family_display
                resp[4],   # prompt
                resp[5],   # response
                resp[6],   # collected_at
            ])

        ws.freeze_panes = "A2"
        self._auto_width(ws)
        ws.column_dimensions["F"].width = 45   # prompt
        ws.column_dimensions["G"].width = 80   # response
