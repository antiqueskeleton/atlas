"""
Tests for VisibilityExcelReport (#33) — the Excel export had zero coverage
despite being directly user-facing (Visibility page's "Export Excel" button).

Fixture shapes match VisibilityAnalytics.summarize_responses()'s real return
dict and VisibilityRepository.list_responses()'s real column order
(id, run_id, provider, model, prompt, response, collected_at, family_display)
rather than guessing, since _sheet_raw_responses indexes that tuple by
position — a real, fragile assumption worth pinning with a test.
"""
from openpyxl import load_workbook

from backend.reports.excel_report import VisibilityExcelReport


def _minimal_analytics(target="Firman"):
    return {
        "target_brand": target,
        "total_responses": 10,
        "target_visibility_score": 40.0,
        "brand_counts": {"Firman": 4, "Honda": 6},
    }


def _full_analytics(target="Firman"):
    return {
        "target_brand": target,
        "total_responses": 10,
        "total_tracked_brands": 2,
        "target_visibility_score": 40.0,
        "provider_visibility_scores": {"openai": 50.0, "anthropic": 30.0},
        "prompt_set_visibility_scores": {"Best Generator Brand": 60.0},
        "first_mention_share": {"Firman": 20.0, "Honda": 80.0},
        "brand_position_counts": {1: {"Firman": 2, "Honda": 8}},
        "brand_position_share": {1: {"Firman": 20.0, "Honda": 80.0}},
        "brand_counts": {"Firman": 4, "Honda": 6},
        "negative_brand_counts": {"Honda": 1},
        "target_negative_rate": 0.0,
        "brand_negative_rate": {"Firman": 0.0, "Honda": 16.7},
        "feature_counts": {"Dual Fuel": 3},
        "feature_brand_counts": {"Dual Fuel": {"Firman": 2, "Honda": 1}},
        "provider_brand_counts": {"openai": {"Firman": 3, "Honda": 2}},
        "channel_counts": {"Amazon": 5},
        "brand_channel_counts": {"Firman": {"Amazon": 1}, "Honda": {"Amazon": 4}},
        "channel_brand_counts": {"Amazon": {"Firman": 1, "Honda": 4}},
        "target_channel_gap": [{
            "channel": "Amazon", "target_count": 1, "top_competitor": "Honda",
            "top_competitor_count": 4, "total_competitor_mentions": 4,
        }],
    }


_STATS = {"runs": 3, "providers": 2, "families": 5}

# Matches VisibilityRepository.list_responses()'s real SELECT column order.
_RAW_RESPONSE = (
    101, "run-1", "openai", "gpt-4.1-mini",
    "best generator brand", "Firman and Honda are both solid choices.",
    "2026-07-01T10:00:00", "Best Generator Brand",
)


def test_generate_produces_valid_xlsx_with_core_sheets(tmp_path):
    out = tmp_path / "report.xlsx"
    rpt = VisibilityExcelReport(_full_analytics(), runs=[], stats=_STATS, target_brand="Firman")
    rpt.generate(str(out))

    assert out.exists() and out.stat().st_size > 0
    wb = load_workbook(str(out))
    assert wb.sheetnames == [
        "Summary", "Brand Analytics", "Provider Scores",
        "Feature Intelligence", "Channel Intelligence",
    ]


def test_raw_responses_sheet_only_appears_when_data_provided(tmp_path):
    out_without = tmp_path / "no_raw.xlsx"
    VisibilityExcelReport(_full_analytics(), [], _STATS, "Firman").generate(str(out_without))
    assert "Raw Responses" not in load_workbook(str(out_without)).sheetnames

    out_with = tmp_path / "with_raw.xlsx"
    VisibilityExcelReport(
        _full_analytics(), [], _STATS, "Firman", raw_responses=[_RAW_RESPONSE],
    ).generate(str(out_with))
    assert "Raw Responses" in load_workbook(str(out_with)).sheetnames


def test_raw_responses_sheet_maps_tuple_columns_correctly(tmp_path):
    """Pins the real VisibilityRepository.list_responses() column order
    against _sheet_raw_responses' positional indexing (resp[0]=id,
    resp[7]=family_display, etc.) — if the repository's SELECT column
    order ever changes, this test catches the silent column-mismatch
    before it ships, rather than a user discovering swapped columns."""
    out = tmp_path / "raw.xlsx"
    VisibilityExcelReport(
        _full_analytics(), [], _STATS, "Firman", raw_responses=[_RAW_RESPONSE],
    ).generate(str(out))

    ws = load_workbook(str(out))["Raw Responses"]
    row = [ws.cell(row=2, column=c).value for c in range(1, 9)]
    assert row == [
        101, "run-1", "openai", "gpt-4.1-mini",
        "Best Generator Brand",  # family_display (resp[7]) lands in column 5
        "best generator brand",  # prompt (resp[4])
        "Firman and Honda are both solid choices.",  # response (resp[5])
        "2026-07-01T10:00:00",  # collected_at (resp[6])
    ]


def test_summary_sheet_kpis_reflect_input_data(tmp_path):
    out = tmp_path / "kpi.xlsx"
    VisibilityExcelReport(_full_analytics(), [], _STATS, "Firman").generate(str(out))
    ws = load_workbook(str(out))["Summary"]

    values = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(6, 13)}
    assert values["Visibility Score"] == "40.0%"
    assert values["Mention Rank"] == "#2 of 2"  # Firman=4 < Honda=6, so 2nd
    assert values["Total Responses"] == "10"
    assert values["Collection Runs"] == "3"


def test_mention_rank_and_brands_tracked_use_total_tracked_not_just_mentioned(tmp_path):
    """
    Regression test: brand_counts only contains brands that were mentioned
    at least once, but Atlas tracks brands that got zero mentions too (e.g.
    a real run: 90 tracked brands, only 70 ever mentioned). The report
    previously computed both Mention Rank's denominator and Brands Tracked
    from len(brand_counts) — silently different from the screen, which
    correctly uses total_tracked_brands (see #48). Deliberately uses a
    fixture where these two numbers differ, unlike _full_analytics()'s
    fixture (used by test_summary_sheet_kpis_reflect_input_data above) where
    they coincidentally match and hid this exact bug.
    """
    out = tmp_path / "tracked.xlsx"
    analytics = _full_analytics()
    analytics["total_tracked_brands"] = 90  # far more than the 2 in brand_counts
    VisibilityExcelReport(analytics, [], _STATS, "Firman").generate(str(out))
    ws = load_workbook(str(out))["Summary"]

    values = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(6, 13)}
    assert values["Mention Rank"] == "#2 of 90"
    assert values["Brands Tracked"] == "90"


def test_target_brand_row_is_highlighted_in_brand_analytics(tmp_path):
    out = tmp_path / "highlight.xlsx"
    VisibilityExcelReport(_full_analytics(), [], _STATS, "Firman").generate(str(out))
    ws = load_workbook(str(out))["Brand Analytics"]

    rows_by_brand = {}
    for r in range(2, ws.max_row + 1):
        brand = ws.cell(r, 1).value
        if brand:
            rows_by_brand[brand] = r

    firman_fill = ws.cell(rows_by_brand["Firman"], 1).fill.fgColor.rgb
    honda_fill = ws.cell(rows_by_brand["Honda"], 1).fill.fgColor.rgb
    assert firman_fill == "00" + VisibilityExcelReport.TARGET
    assert honda_fill != firman_fill


def test_generate_does_not_crash_with_minimal_analytics(tmp_path):
    """Most fields are optional (.get(..., {})) — confirm the report still
    generates cleanly from a near-empty analytics dict, e.g. a brand-new
    installation with only a couple of collected responses so far."""
    out = tmp_path / "minimal.xlsx"
    VisibilityExcelReport(_minimal_analytics(), [], {}, "Firman").generate(str(out))
    assert out.exists()

    wb = load_workbook(str(out))
    # Sheets with no underlying data still get created but stay empty —
    # this is real, current behavior (create_sheet() runs before the
    # early-return guard in each sheet method), not something this test
    # is asserting SHOULD happen, just confirming it doesn't crash.
    assert "Feature Intelligence" in wb.sheetnames
    assert wb["Feature Intelligence"].max_row == 1  # no rows written


def test_generate_does_not_crash_with_no_target_brand_mentions(tmp_path):
    """Target brand tracked but never mentioned in any response — rank
    lookup (`next(... if b == tb)`) must not raise when tb isn't in
    brand_counts at all."""
    analytics = _full_analytics(target="Generac")  # not in brand_counts
    out = tmp_path / "no_mentions.xlsx"
    VisibilityExcelReport(analytics, [], _STATS, "Generac").generate(str(out))

    ws = load_workbook(str(out))["Summary"]
    values = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(6, 13)}
    assert values["Mention Rank"] == "—"
